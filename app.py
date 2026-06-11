import os
import re
import shutil
import hashlib
import datetime
import traceback
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from dotenv import load_dotenv

from modules.pdf_to_image import convert_pdf_to_images
from modules.image_cleaner import clean_image
from modules.claude_extractor import extract_attendance_from_image
from modules.excel_generator import generate_excel, parse_existing_sheet
from modules.attendance_processor import process_attendance

load_dotenv()

app = Flask(__name__)
# Enable CORS for all routes (necessary for Vercel -> Render communication)
CORS(app)

# Ensure folders exist
os.makedirs("input_pdfs", exist_ok=True)
os.makedirs("processed_pdfs", exist_ok=True)
os.makedirs("output_excels", exist_ok=True)
os.makedirs("temp_images", exist_ok=True)

CONTRACTOR_MAPPING = {
    "Shri Shyam": {"id": "shyam", "color": "var(--navy)"},
    "Ayush": {"id": "ayush", "color": "var(--teal-brand)"},
    "Shri Laxmi": {"id": "laxmi", "color": "var(--emerald-brand)"},
    "Ansh": {"id": "ansh", "color": "var(--warning)"},
    "Shri Radha": {"id": "radha", "color": "var(--slate-cool)"},
    "Jaishil Sulphur and Chemical Industries": {"id": "jaishil", "color": "var(--navy-deep)"}
}

REV_CONTRACTOR_MAPPING = {v["id"]: k for k, v in CONTRACTOR_MAPPING.items()}


def get_employee_id(contractor_name, shift, name):
    unique_str = f"{contractor_name}_{shift}_{name}"
    return "e_" + hashlib.md5(unique_str.encode('utf-8')).hexdigest()[:8]


def run_pipeline_on_pdf(pdf_path, pdf_file):
    temp_folder = "temp_images"
    cleaned_folder = "temp_images/cleaned"
    
    # 1. Parse date from filename
    pdf_date = None
    match = re.search(r"([A-Za-z]+)\s+(\d{1,2}),\s*(\d{4})", pdf_file)
    if match:
        month_name = match.group(1)
        day = int(match.group(2))
        year = int(match.group(3))
        month_number = datetime.datetime.strptime(month_name, "%B").month
        pdf_date = f"{year}-{month_number:02d}-{day:02d}"
        month_key = f"{year}_{month_number:02d}"
    else:
        # fallback to current date
        now = datetime.datetime.now()
        pdf_date = now.strftime("%Y-%m-%d")
        month_key = now.strftime("%Y_%m")
        
    # 2. PDF to images
    image_paths = convert_pdf_to_images(pdf_path, temp_folder)
    
    # 3. Clean images
    cleaned_images = []
    for image_path in image_paths:
        cleaned_image = clean_image(image_path, cleaned_folder)
        cleaned_images.append(cleaned_image)
        
    # 4. Extract attendance from each page
    all_results = []
    for index, image_path in enumerate(cleaned_images, start=1):
        extracted_data = extract_attendance_from_image(image_path)
        
        # Use extracted date (fallback to pdf_date)
        for row in extracted_data:
            extracted_date = row.get("date")
            if not extracted_date or not re.match(r"^\d{4}-\d{2}-\d{2}$", str(extracted_date)):
                row["date"] = pdf_date
                
        all_results.extend(extracted_data)
        
    # 5. Group results by month
    primary_excel_file = f"output_excels/attendance_{month_key}.xlsx"
    if not all_results:
        generate_excel({}, primary_excel_file)
    else:
        from collections import defaultdict
        results_by_month = defaultdict(list)
        for row in all_results:
            row_date = row.get("date")
            if row_date and re.match(r"^\d{4}-\d{2}-\d{2}$", str(row_date)):
                parts = row_date.split("-")
                row_month_key = f"{parts[0]}_{parts[1]}"
            else:
                row_month_key = month_key
            results_by_month[row_month_key].append(row)
            
        for row_month_key, group_results in results_by_month.items():
            structured_data = process_attendance(group_results)
            excel_file = f"output_excels/attendance_{row_month_key}.xlsx"
            generate_excel(structured_data, excel_file)
            
    # 6. Move PDF safely
    if os.path.exists(pdf_path):
        dest_path = os.path.join("processed_pdfs", pdf_file)
        if os.path.exists(dest_path):
            try:
                os.remove(dest_path)
            except Exception as e:
                print(f"Warning: Could not remove existing file {dest_path}: {e}")
        shutil.move(pdf_path, dest_path)
        
    return len(all_results), primary_excel_file


@app.route("/api/upload", methods=["POST"])
def upload_file():
    if 'file' not in request.files:
        return jsonify({"error": "No file part in the request"}), 400
        
    file = request.files['file']
    if file.filename == '':
        return jsonify({"error": "No file selected"}), 400
        
    if not file.filename.lower().endswith('.pdf'):
        return jsonify({"error": "Only PDF files are allowed"}), 400
        
    pdf_path = os.path.join("input_pdfs", file.filename)
    file.save(pdf_path)
    
    try:
        records_count, excel_file = run_pipeline_on_pdf(pdf_path, file.filename)
        return jsonify({
            "ok": True,
            "filename": file.filename,
            "records": records_count,
            "excel": os.path.basename(excel_file)
        })
    except Exception as e:
        # Clean up input file if pipeline fails
        traceback.print_exc()
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        return jsonify({"error": str(e)}), 500


@app.route("/api/attendance", methods=["GET"])
def get_attendance():
    month_key = request.args.get("month")  # format: YYYY_MM (e.g. 2026_05)
    if not month_key:
        # Fallback to current month if not provided
        now = datetime.datetime.now()
        month_key = now.strftime("%Y_%m")
        
    excel_file = f"output_excels/attendance_{month_key}.xlsx"
    
    contractors_list = []
    employees_list = []
    attendance_map = {}
    
    # Initialize contractors list
    for c_name, c_info in CONTRACTOR_MAPPING.items():
        contractors_list.append({
            "id": c_info["id"],
            "name": c_name,
            "employees": 0,
            "color": c_info["color"]
        })
        
    if os.path.exists(excel_file):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Retrieve the full contractor name
                title_val = ws.cell(1, 1).value
                contractor_name = sheet_name
                if title_val and title_val.endswith(" ATTENDANCE"):
                    contractor_name = title_val[:-11].strip()
                    
                c_info = CONTRACTOR_MAPPING.get(contractor_name)
                c_id = c_info["id"] if c_info else contractor_name.lower().replace(" ", "_")
                
                parsed_sheet = parse_existing_sheet(ws)
                
                # Set employee count
                cnt = sum(len(parsed_sheet.get(s, {})) for s in ["DAY", "NIGHT"])
                for c in contractors_list:
                    if c["id"] == c_id:
                        c["employees"] = cnt
                        
                for shift_name in ["DAY", "NIGHT"]:
                    emps = parsed_sheet.get(shift_name, {})
                    for emp_name, att_data in emps.items():
                        emp_id = get_employee_id(contractor_name, shift_name, emp_name)
                        employees_list.append({
                            "id": emp_id,
                            "name": emp_name,
                            "contractorId": c_id,
                            "shift": shift_name
                        })
                        
                        attendance_map[emp_id] = {}
                        for day_num, day_vals in att_data.items():
                            attendance_map[emp_id][day_num] = {
                                "in": day_vals.get("in", ""),
                                "out": day_vals.get("out", ""),
                                "ot": day_vals.get("ot", "")
                            }
        except Exception as e:
            traceback.print_exc()
            return jsonify({"error": f"Error parsing excel file: {str(e)}"}), 500
            
    return jsonify({
        "month": month_key,
        "contractors": contractors_list,
        "employees": employees_list,
        "attendance": attendance_map
    })


@app.route("/api/attendance/save", methods=["POST", "PUT"])
def save_attendance():
    payload = request.json
    if not payload:
        return jsonify({"error": "Missing payload"}), 400
        
    month_key = payload.get("month")
    if not month_key:
        return jsonify({"error": "Missing month in payload"}), 400
        
    attendance_data = payload.get("attendance", {})
    employees_list = payload.get("employees", [])
    contractors_list = payload.get("contractors", [])
    
    # Reconstruct structured_data
    structured_data = {}
    
    for emp in employees_list:
        emp_id = emp["id"]
        emp_name = emp["name"]
        shift = emp["shift"]
        c_id = emp["contractorId"]
        
        contractor_name = "UNKNOWN"
        for c in contractors_list:
            if c["id"] == c_id:
                contractor_name = c["name"]
                break
        if contractor_name == "UNKNOWN":
            contractor_name = REV_CONTRACTOR_MAPPING.get(c_id, c_id)
            
        if contractor_name not in structured_data:
            structured_data[contractor_name] = {"DAY": {}, "NIGHT": {}}
            
        emp_att = attendance_data.get(emp_id, {})
        
        daily_data = {}
        for day_str, vals in emp_att.items():
            try:
                day_num = int(day_str)
                if vals.get("in") or vals.get("out") or vals.get("ot"):
                    daily_data[day_num] = {
                        "in": vals.get("in", ""),
                        "out": vals.get("out", ""),
                        "ot": vals.get("ot", "")
                    }
            except ValueError:
                continue
                
        if daily_data:
            structured_data[contractor_name][shift][emp_name] = daily_data
            
    excel_file = f"output_excels/attendance_{month_key}.xlsx"
    try:
        generate_excel(structured_data, excel_file)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"error": f"Failed to save excel: {str(e)}"}), 500


@app.route("/api/files", methods=["GET"])
def get_files():
    files = []
    folder = "output_excels"
    if os.path.exists(folder):
        for fname in os.listdir(folder):
            if fname.lower().endswith(".xlsx") and fname.startswith("attendance_"):
                fpath = os.path.join(folder, fname)
                stat = os.stat(fpath)
                mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
                
                # Count actual records/rows written
                record_count = 0
                try:
                    from openpyxl import load_workbook
                    wb = load_workbook(fpath, data_only=True)
                    for sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        parsed_sheet = parse_existing_sheet(ws)
                        for shift, emps in parsed_sheet.items():
                            for emp, att_data in emps.items():
                                record_count += len(att_data)
                except Exception as e:
                    print(f"Error counting records for {fname}: {e}")
                    
                files.append({
                    "name": fname,
                    "size": f"{stat.st_size // 1024} KB",
                    "generatedAt": mtime.strftime("%b %d, %Y · %H:%M"),
                    "records": record_count
                })
    files.sort(key=lambda x: x["name"], reverse=True)
    return jsonify(files)


@app.route("/api/download/<filename>", methods=["GET"])
def download_file(filename):
    directory = os.path.abspath("output_excels")
    return send_from_directory(directory, filename, as_attachment=True)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
