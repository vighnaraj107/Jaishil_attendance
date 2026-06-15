import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from modules.styles import *


def format_time_value(val):
    if val is None:
        return ""
    if isinstance(val, datetime.time):
        return val.strftime("%H:%M")
    if isinstance(val, datetime.datetime):
        return val.time().strftime("%H:%M")
    val_str = str(val).strip()
    if val_str.upper() in ["NONE", ""]:
        return ""
    return val_str


def parse_existing_sheet(ws):
    shift_data = {
        "DAY": {},
        "NIGHT": {}
    }
    current_shift = "DAY"
    
    row = 5
    max_row = ws.max_row
    while row <= max_row:
        val_1 = ws.cell(row, 1).value
        val_2 = ws.cell(row, 2).value
        
        # Check if shift header
        if val_1 and "SHIFT" in str(val_1).upper():
            if "DAY" in str(val_1).upper():
                current_shift = "DAY"
            elif "NIGHT" in str(val_1).upper():
                current_shift = "NIGHT"
            row += 1
            continue
        
        # Check if employee row
        if val_1 and val_2 == "IN":
            employee_name = val_1
            employee_data = {}
            
            # Determine sheet layout format dynamically
            has_total_row = (ws.cell(row + 2, 2).value == "TOTAL")
            ot_offset = 3 if has_total_row else 2
            total_offset = 2 if has_total_row else None
            
            for day in range(1, 32):
                col = day + 2
                in_time = ws.cell(row, col).value
                out_time = ws.cell(row + 1, col).value
                ot_time = ws.cell(row + ot_offset, col).value
                total_time = ws.cell(row + total_offset, col).value if total_offset is not None else None
                
                # Only save if we have some data for this day
                if in_time is not None or out_time is not None or ot_time is not None or total_time is not None:
                    formatted_in = format_time_value(in_time)
                    formatted_out = format_time_value(out_time)
                    formatted_ot = format_time_value(ot_time)
                    
                    if has_total_row:
                        formatted_total = format_time_value(total_time)
                    else:
                        from modules.calculations import calculate_work_hours
                        formatted_total = calculate_work_hours(formatted_in, formatted_out) if (formatted_in and formatted_out) else ""
                    
                    employee_data[day] = {
                        "in": formatted_in,
                        "out": formatted_out,
                        "work_hours": formatted_total,
                        "ot": formatted_ot
                    }
            shift_data[current_shift][employee_name] = employee_data
            row += 5 if has_total_row else 4  # skip IN, OUT, (TOTAL), OT and the spacer row
            continue
            
        row += 1
        
    return shift_data


def merge_attendance_data(existing_data, new_data):
    combined = {}
    
    # Identify which (contractor, day) pairs are being updated in the new data
    contractor_days_to_clear = set()
    for contractor, shifts in new_data.items():
        for shift, employees in shifts.items():
            for employee, attendance in employees.items():
                for day in attendance.keys():
                    contractor_days_to_clear.add((contractor, day))
                    
    # 1. Copy all existing data, except for the (contractor, day) pairs being updated
    for contractor, shifts in existing_data.items():
        combined[contractor] = {"DAY": {}, "NIGHT": {}}
        for shift, employees in shifts.items():
            for employee, attendance in employees.items():
                # Filter out days that are being cleared for this contractor
                cleaned_attendance = {
                    day: data for day, data in attendance.items()
                    if (contractor, day) not in contractor_days_to_clear
                }
                combined[contractor][shift][employee] = cleaned_attendance
                
    # 2. Merge new data
    for contractor, shifts in new_data.items():
        if contractor not in combined:
            combined[contractor] = {"DAY": {}, "NIGHT": {}}
        for shift, employees in shifts.items():
            for employee, attendance in employees.items():
                if employee not in combined[contractor][shift]:
                    combined[contractor][shift][employee] = {}
                for day, data in attendance.items():
                    combined[contractor][shift][employee][day] = data
                    
    return combined


def generate_excel(structured_data, output_path):

    existing_data = {}
    
    # 1. Load existing data if file exists
    if os.path.exists(output_path):
        print(
            f"\nExisting workbook found: {output_path}"
            "\nLoading existing data to merge...\n"
        )
        try:
            wb = load_workbook(output_path)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Extract contractor name from title cell A1
                title_val = ws.cell(1, 1).value
                contractor_name = sheet_name
                if title_val and title_val.endswith(" ATTENDANCE"):
                    contractor_name = title_val[:-11].strip()
                
                existing_data[contractor_name] = parse_existing_sheet(ws)
        except Exception as e:
            print(f"Warning: Failed to load/parse existing workbook: {e}. Writing from scratch.")
            existing_data = {}
            
    # 2. Merge existing data with new data
    merged_data = merge_attendance_data(existing_data, structured_data)
    
    # 3. Create a fresh workbook to write the merged data
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    
    # 4. Write merged data
    for contractor, shifts in merged_data.items():
        print(f"Processing sheet: {contractor}")
        
        sheet_name = contractor[:30]
        ws = wb.create_sheet(title=sheet_name)
        
        # ──────────────────────────────────────────────────────────────────────
        # TITLE ROW
        # ──────────────────────────────────────────────────────────────────────
        ws.merge_cells("A1:AJ1")
        title = ws["A1"]
        title.value = f"{contractor} ATTENDANCE"
        title.fill = header_fill
        title.font = Font(bold=True, color="FFFFFF", size=16)
        title.alignment = center
        
        # ──────────────────────────────────────────────────────────────────────
        # HEADER ROW (row 3)
        # ──────────────────────────────────────────────────────────────────────
        header_row = 3
        ws.cell(header_row, 1).value = "Employee Name"
        ws.cell(header_row, 2).value = "Type"
        
        for day in range(1, 32):
            ws.cell(header_row, day + 2).value = day
            
        summary_col = 35
        ws.cell(header_row, summary_col).value = "Total Days"
        ws.cell(header_row, summary_col + 1).value = "Total OT"
        ws.cell(header_row, summary_col + 2).value = "OT Days"
        
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center
            
        # ──────────────────────────────────────────────────────────────────────
        # SHIFT SECTIONS
        # ──────────────────────────────────────────────────────────────────────
        current_row = 5
        
        for shift_name in ["DAY", "NIGHT"]:
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=36
            )
            shift_cell = ws.cell(current_row, 1)
            shift_cell.value = f"{shift_name} SHIFT"
            shift_cell.fill = section_fill
            shift_cell.font = white_bold_font
            shift_cell.alignment = center
            
            current_row += 1
            
            employees = shifts.get(shift_name, {})
            
            # ──────────────────────────────────────────────────────────────────
            # EMPLOYEES
            # ──────────────────────────────────────────────────────────────────
            for employee, attendance in employees.items():
                start_row = current_row
                
                ws.merge_cells(
                    start_row=start_row,
                    start_column=1,
                    end_row=start_row + 3,
                    end_column=1
                )
                
                ws.cell(start_row, 1).value = employee
                ws.cell(start_row, 2).value = "IN"
                ws.cell(start_row + 1, 2).value = "OUT"
                ws.cell(start_row + 2, 2).value = "TOTAL"
                ws.cell(start_row + 3, 2).value = "OT"
                
                # ──────────────────────────────────────────────────────────────
                # WRITE DAILY DATA
                # ──────────────────────────────────────────────────────────────
                for day in range(1, 32):
                    col = day + 2
                    if day in attendance:
                        data = attendance[day]
                        in_val = data.get("in", "")
                        out_val = data.get("out", "")
                        ot_val = data.get("ot", "")
                        
                        # Dynamically calculate work_hours if it's missing but we have in/out
                        work_hours_val = data.get("work_hours", "")
                        if in_val and out_val and not work_hours_val:
                            from modules.calculations import calculate_work_hours
                            work_hours_val = calculate_work_hours(in_val, out_val)
                            
                        ws.cell(start_row, col).value = in_val
                        ws.cell(start_row + 1, col).value = out_val
                        ws.cell(start_row + 2, col).value = work_hours_val
                        ws.cell(start_row + 3, col).value = ot_val
                        
                # ──────────────────────────────────────────────────────────────
                # RECALCULATE SUMMARY FROM THE FULL SHEET
                # ──────────────────────────────────────────────────────────────
                total_days = 0
                total_ot_minutes = 0
                
                for day in range(1, 32):
                    col = day + 2
                    in_value = ws.cell(start_row, col).value
                    ot_value = ws.cell(start_row + 3, col).value
                    
                    if in_value:
                        total_days += 1
                        
                    if ot_value:
                        try:
                            hrs, mins = map(int, str(ot_value).split(":"))
                            total_ot_minutes += hrs * 60 + mins
                        except:
                            pass
                            
                ot_hours = total_ot_minutes // 60
                ot_mins = total_ot_minutes % 60
                total_ot = f"{ot_hours:02}:{ot_mins:02}"
                ot_days = round(total_ot_minutes / 480, 2)
                
                ws.cell(start_row, summary_col).value = total_days
                ws.cell(start_row, summary_col + 1).value = total_ot
                ws.cell(start_row, summary_col + 2).value = ot_days
                
                current_row += 5
                
        # ──────────────────────────────────────────────────────────────────────
        # STYLING
        # ──────────────────────────────────────────────────────────────────────
        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = center
                
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 12
        
        for col in range(3, 40):
            ws.column_dimensions[get_column_letter(col)].width = 12
            
    # ──────────────────────────────────────────────────────────────────────────
    # SAVE
    # ──────────────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f"\nExcel Saved: {output_path}\n")
