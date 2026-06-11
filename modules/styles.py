from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import Alignment

header_fill = PatternFill(
    start_color="1F4E78",
    end_color="1F4E78",
    fill_type="solid"
)

section_fill = PatternFill(
    start_color="17365D",
    end_color="17365D",
    fill_type="solid"
)

green_fill = PatternFill(
    start_color="C6EFCE",
    end_color="C6EFCE",
    fill_type="solid"
)

header_font = Font(
    bold=True,
    color="FFFFFF",
    size=11
)

white_bold_font = Font(
    bold=True,
    color="FFFFFF"
)

thin = Side(style='thin')

border = Border(
    left=thin,
    right=thin,
    top=thin,
    bottom=thin
)

center = Alignment(
    horizontal="center",
    vertical="center"
)